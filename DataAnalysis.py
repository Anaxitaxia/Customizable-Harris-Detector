import openpyxl
from openpyxl.styles import PatternFill


def save_data(img_name, opportunity_lst, parametr_lst, result_corners_amount_lst, ttime):
    str_ind = img_name.rfind('/') + 1
    wb = openpyxl.load_workbook("Results.xlsx")
    index = 3
    sheet = wb.active
    while sheet.cell(row=2, column=index).value:
        index += 1
    sheet.cell(row=2, column=index).value = img_name[str_ind:len(img_name):1]
    for row_ind in range(3, 16):
        sheet.cell(row=row_ind, column=index).value = opportunity_lst[row_ind - 3]
    for row_ind in range(16, 24):
        sheet.cell(row=row_ind, column=index).value = parametr_lst[row_ind - 16]
    for row_ind in range(24, 31):
        sheet.cell(row=row_ind, column=index).value = result_corners_amount_lst[row_ind - 24]
    sheet.cell(row=37, column=index).value = ttime
    wb.save("Results.xlsx")


def analyze_data(answer, right, untruly, wrong, extra):
    wb = openpyxl.load_workbook("Results.xlsx")
    sheet = wb.active
    col_ind = 3
    green_fill = PatternFill('solid', fgColor='32CD32')
    red_fill = PatternFill('solid', fgColor='FF0000')
    ye_green_fill = PatternFill(fill_type='solid', fgColor='ADFF2F')
    yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')
    orrange_fill = PatternFill(fill_type='solid', fgColor='FF8C00')

    while sheet.cell(row=2, column=col_ind).value:
        col_ind += 1
    col_ind -= 1
    if (right == answer) and (untruly == 0) and (extra == 0):
        for row_ind in range(2, 38):
            sheet.cell(row=row_ind, column=col_ind).fill = green_fill
    elif (right + untruly == answer) and (extra == 0):
        for row_ind in range(2, 38):
            sheet.cell(row=row_ind, column=col_ind).fill = ye_green_fill
    elif (right + untruly == answer) and (wrong == 0):
        for row_ind in range(2, 38):
            sheet.cell(row=row_ind, column=col_ind).fill = yellow_fill
    elif (right + untruly == answer) and (wrong < 10):
        for row_ind in range(2, 38):
            sheet.cell(row=row_ind, column=col_ind).fill = orrange_fill
    elif wrong > 10:
        for row_ind in range(2, 38):
            sheet.cell(row=row_ind, column=col_ind).fill = red_fill
    wb.save("Results.xlsx")


def count_corners(answer_corners, corners):
    list_question = []
    list_claster = []
    ch_right = 0
    s_right = []
    ch_untruly = 0
    s_untruly = []
    ch_claster = 0
    ch_wrong = 0
    s_wrong = []
    ch_answer = len(answer_corners)
    for i in range(len(corners)):
        flag = False
        for c in answer_corners:
            if (abs(corners[i][0] - c[0]) < 1) and (abs(corners[i][1] - c[1]) < 1):
                ch_right += 1
                s_right.append(corners[i])
                flag = True
            else:
                fl = False
                for n in (-1, 0, 1):
                    for m in (-1, 0, 1):
                        corn1 = corners[i][0] + n
                        corn2 = corners[i][1] + m
                        if (abs(corn1 - c[0]) < 1) and (abs(corn2 - c[1]) < 1):
                            ch_claster += 1
                            list_question.append([c[0], c[1]])
                            list_claster.append(corners[i])
                            fl = True
                            flag = True
                            break
                    if fl:
                        break
        if not flag:
            ch_wrong += 1
            s_wrong.append(corners[i])
    if ch_right == ch_answer:
        ch_untruly = 0
    else:
        for i in range(len(list_question)):
            if not (list_question[i] in s_right) and not (list_question[i] in list_question[i + 1:len(list_question)]):
                s_untruly.append(list_claster[i])
                ch_claster -= 1
                ch_untruly += 1
    for corn in s_untruly:
        list_claster.remove(corn)
    ch_extra = ch_wrong + ch_claster
    list_ch = [ch_answer, ch_right, ch_untruly, ch_claster, ch_wrong, ch_extra]
    list_corners = [s_right, s_untruly, list_claster, s_wrong]
    return list_ch, list_corners
