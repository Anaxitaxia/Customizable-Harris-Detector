from PIL import Image
import numpy as np
import HD
import cv2 as cv
import time
import os
import concurrent.futures as cf
import openpyxl


def print_main_menu():
    print('1. Весовая функция (по умолчанию --- Гауссова).')
    print('2. Размер окна весовой функции (по умолчанию --- 3).')
    print('3. Параметр сигма функции Гаусса (по умолчанию --- 1).')
    print('4. Мера отклика (по умолчанию --- мера Харриса)')
    print('5. Параметр k меры Харриса (по умолчанию --- 0.04)')
    print('6. Подавление немаксимумов (по умолчанию --- включено)')
    print('7. Размер окна для поиска локальных максимумов (по умолчанию --- 3)')
    print('8. Режим log-log scaled detector (по умолчанию --- выключено)')
    print('9. Параметр b у log-log scaled detector (по умолчанию --- 0.1)')
    print('10. Политика отсечения углов (по умолчанию --- адаптивное пороговое значение)')
    print('11. Пороговое значение p (по умолчанию --- 0.005)')
    print('12. Режим предварительного отсечения точек (по умолчанию --- выключено)')
    print('13. Пороговое значение для предварительного отсечения точек (по умолчанию --- 200)')
    print('14. Режим выделения значимого региона (по умолчанию --- выключено)')
    print('15. Выход')


def find_im_names(path_name):
    def find_file_names(im_name):
        return path_name + str(im_name)
    im_names = list(os.walk(path_name))[0][2]
    return list(map(find_file_names, im_names))


def find_corners(im_name):
    imag = Image.open(im_name).convert('L')
    imag.load()
    im = np.array(imag)
    all_results = []
    for k1 in k_list:
        for p1 in p_list:
            for b1 in b_list:
                for non_max_size1 in non_max_size_list:
                    for cut_th1 in cut_th_list:
                        t1 = time.time()
                        harris_det = HD.HarrisDetector(im, weight_func, sigma, weight_wind_size, resp, k1,
                                                       th_politic, p1, non_max_fl, non_max_size1, log_log,
                                                       b1, cut_fl, cut_th1, significant_fl)
                        t2 = time.time()
                        y_coord, x_coord, corner_map = harris_det.find_corners()
                        all_results.append([im_name, x_coord, y_coord, t2 - t1, weight_func, resp, k1, p1, non_max_fl,
                                            non_max_size1, log_log, b1, cut_fl, cut_th1, significant_fl])
    return all_results


def save_results():
    try:
        wb = openpyxl.load_workbook(result_file)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1, value='Имя')
        sheet.cell(row=1, column=2, value='Координата x угла')
        sheet.cell(row=1, column=3, value='Координата y угла')
        sheet.cell(row=1, column=4, value='Время на поиск углов, сек')
        sheet.cell(row=1, column=5, value='Весовая функция')
        sheet.cell(row=1, column=6, value='Мера отклика')
        sheet.cell(row=1, column=7, value='k')
        sheet.cell(row=1, column=8, value='p')
        sheet.cell(row=1, column=9, value='Метод подавления немаксимумов')
        sheet.cell(row=1, column=10, value='Размер окна для подавления немаксимумов')
        sheet.cell(row=1, column=11, value='log-log scaled')
        sheet.cell(row=1, column=12, value='b')
        sheet.cell(row=1, column=13, value='Предварительное отсечение точек')
        sheet.cell(row=1, column=14, value='Порог предварительного отсечения')
        sheet.cell(row=1, column=15, value='Выделение значимого региона')
    ind = 2
    while sheet.cell(row=ind, column=2).value:
        ind += 1
    for item in f:
        sheet.cell(row=ind, column=1, value=item[0][0])
        sheet.cell(row=ind, column=4, value=item[0][3])
        sheet.cell(row=ind, column=5, value=item[0][4])
        sheet.cell(row=ind, column=6, value=item[0][5])
        sheet.cell(row=ind, column=7, value=item[0][6])
        sheet.cell(row=ind, column=8, value=item[0][7])
        sheet.cell(row=ind, column=9, value=item[0][8])
        sheet.cell(row=ind, column=10, value=item[0][9])
        sheet.cell(row=ind, column=11, value=item[0][10])
        sheet.cell(row=ind, column=12, value=item[0][11])
        sheet.cell(row=ind, column=13, value=item[0][12])
        sheet.cell(row=ind, column=14, value=item[0][13])
        sheet.cell(row=ind, column=15, value=item[0][14])
        for i in range(item[0][1].shape[0]):
            sheet.cell(row=ind, column=2, value=item[0][1][i])
            sheet.cell(row=ind, column=3, value=item[0][2][i])
            ind += 1
    wb.save(result_file)


# инициализация переменных (параметров и флагов)
weight_func = 'gauss'
resp = 'harris'
th_politic = 'adapt'
non_max_fl = True
log_log = False
cut_fl = False
significant_fl = False
sigma = 1
weight_wind_size = 3
k = 0.04
p = 0.005
non_max_size = 3
b = 0.1
cut_th = 200

# выбор режима работы
print('Выберите режим работы:')
print('    1. Проверить работу детектора на одном изображении.')
print('    2. Исследовать работу детектора.')
print('    3. Выход.')
mode = int(input())
if mode == 1:
    # предобработка изображения
    print('Введите имя изображения в формате C://путь/к/файлу/название.расширение')
    filename = input()
    image = Image.open(filename).convert('L')
    image.load()
    img = np.array(image)
    # настройка параметров
    print('Настройте параметры детектора')
    punct = 1
    while punct != 15:
        print_main_menu()
        punct = int(input())
        if punct == 1:
            print('    1. Гауссова')
            print('    2. B-сплайн')
            weight_func = int(input())
            if weight_func == 1:
                weight_func = 'gauss'
            elif weight_func == 2:
                weight_func = 'b-spline'
        elif punct == 2:
            print('    Введите размер окна: ')
            weight_wind_size = float(input())
        elif punct == 3:
            print('    Введите sigma: ')
            sigma = float(input())
        elif punct == 4:
            print('    1. Мера отклика Харриса')
            print('    2. Мера отклика Фёрстнера (обратите внимание, что в этом случае пороговое значение должно быть '
                  'сильно увеличено)')
            resp = int(input())
            if resp == 1:
                resp = 'harris'
            elif resp == 2:
                resp = 'forstner'
        elif punct == 5:
            print('    Введите параметр k: ')
            k = float(input())
        elif punct == 6:
            print('    1. Включить подавление')
            print('    2. Выключить подавление')
            non_max_fl = int(input()) - 1
            non_max_fl = bool(non_max_fl)
        elif punct == 7:
            print('    Введите размер окна: ')
            non_max_size = int(input())
        elif punct == 8:
            print('    1. Включить log-log scaled')
            print('    2. Выключить log-log scaled')
            log_fl = int(input()) - 1
            log_fl = bool(log_fl)
        elif punct == 9:
            print('    Введите параметр b: ')
            b = float(input())
        elif punct == 10:
            print('    1. Адаптивное пороговое значение')
            print('    2. Порог')
            th_politic = int(input())
            if th_politic == 1:
                th_politic = 'adapt'
            else:
                th_politic = 'no adapt'
        elif punct == 11:
            print('    Введите пороговое значение p: ')
            p = float(input())
        elif punct == 12:
            print('    1. Включить предварительное отсечение')
            print('    2. Выключить предварительное отсечение')
            cut_fl = int(input()) - 1
            cut_fl = bool(cut_fl)
        elif punct == 13:
            print('    Введите пороговое значение для предварительного отсечения точек: ')
            cut_th = int(input())
        elif punct == 14:
            print('    1. Включить выделение значимого региона')
            print('    2. Выключить выделение значимого региона')
            significant_fl = int(input()) - 1
            significant_fl = bool(significant_fl)
        else:
            if punct != 15:
                print('Ошибка. Повторите ввод.')
    # поиск углов
    HarDet = HD.HarrisDetector(img, weight_func, sigma, weight_wind_size, resp, k, th_politic, p, non_max_fl,
                               non_max_size, log_log, b, cut_fl, cut_th, significant_fl)
    tstart = time.time()
    y, x, c_map = HarDet.find_corners()
    tfinish = time.time()
    print('Время на поиск углов, сек.: ', tfinish - tstart)
    # отрисовка углов
    for index in range(x.shape[0]):
        cv.circle(img, (x[index], y[index]), 5, 0, 2)
    image = Image.fromarray(img.astype('uint8'), 'L')
    image.show()
elif mode == 2:
    # инициализация переменных для исследования детектора
    k_start = 0.04
    k_finish = 0.04
    k_step = 0
    non_max_size_start = 3
    non_max_size_finish = 3
    non_max_size_step = 0
    b_start = 0.1
    b_finish = 0.1
    b_step = 0
    p_start = 0.005
    p_finish = 0.005
    p_step = 0
    cut_th_start = 200
    cut_th_finish = 200
    cut_th_step = 0
    print('Введите пути к изображениям. Чтобы закончить ввод, нажмите n/N')
    symb = 'y'
    path_list = []
    # ввод путей
    while symb.lower() != 'n':
        print('Введите пути к файлам в формате C://путь/к/файлу/')
        path = input()
        path_list.append(path)
        print('Продолжить ввод?')
        symb = input()
    # настройка параметров
    print('Настройте параметры детектора')
    punct = 1
    while punct != 15:
        print_main_menu()
        punct = int(input())
        if punct == 1:
            print('    1. Гауссова')
            print('    2. B-сплайн')
            weight_func = int(input())
            if weight_func == 1:
                weight_func = 'gauss'
            elif weight_func == 2:
                weight_func = 'b-spline'
        elif punct == 2:
            print('    Введите размер окна: ')
            weight_wind_size = float(input())
        elif punct == 3:
            print('    Введите sigma: ')
            sigma = float(input())
        elif punct == 4:
            print('    1. Мера отклика Харриса')
            print('    2. Мера отклика Фёрстнера (обратите внимание, что в этом случае пороговое значение должно быть '
                  'сильно увеличено)')
            resp = int(input())
            if resp == 1:
                resp = 'harris'
            elif resp == 2:
                resp = 'forstner'
        elif punct == 5:
            print('    Введите начальный параметр k: ')
            k_start = float(input())
            print('    Введите конечный параметр k: ')
            k_finish = float(input())
            print('    Введите шаг параметра k: ')
            k_step = float(input())
        elif punct == 6:
            print('    1. Включить подавление')
            print('    2. Выключить подавление')
            non_max_fl = int(input()) - 1
            non_max_fl = bool(non_max_fl)
        elif punct == 7:
            print('    Введите начальный размер окна: ')
            non_max_size_start = int(input())
            print('    Введите конечный размер окна (учтите, что шаг равен единице): ')
            non_max_size_finish = int(input())
            print('    Введите шаг размера окна')
            non_max_size_step = int(input())
        elif punct == 8:
            print('    1. Включить log-log scaled')
            print('    2. Выключить log-log scaled')
            log_fl = int(input()) - 1
            log_fl = bool(log_fl)
        elif punct == 9:
            print('    Введите начальный параметр b: ')
            b_start = float(input())
            print('    Введите конечный параметр b: ')
            b_finish = float(input())
            print('    Введите шаг параметра b: ')
            b_step = float(input())
        elif punct == 10:
            print('    1. Адаптивное пороговое значение')
            print('    2. Порог')
            th_politic = int(input())
            if th_politic == 1:
                th_politic = 'adapt'
            else:
                th_politic = 'no adapt'
        elif punct == 11:
            print('    Введите начальное пороговое значение p: ')
            p_start = float(input())
            print('    Введите конечное пороговое значение p: ')
            p_finish = float(input())
            print('    Введите шаг порогового значения p: ')
            p_step = float(input())
        elif punct == 12:
            print('    1. Включить предварительное отсечение')
            print('    2. Выключить предварительное отсечение')
            cut_fl = int(input()) - 1
            cut_fl = bool(cut_fl)
        elif punct == 13:
            print('    Введите начальное пороговое значение для предварительного отсечения точек: ')
            cut_th_start = int(input())
            print('    Введите конечное пороговое значение для предварительного отсечения точек: ')
            cut_th_finish = int(input())
            print('    Введите шаг порогового значения для предварительного отсечения точек: ')
            cut_th_step = int(input())
        elif punct == 14:
            print('    1. Включить выделение значимого региона')
            print('    2. Выключить выделение значимого региона')
            significant_fl = int(input()) - 1
            significant_fl = bool(significant_fl)
        else:
            if punct != 15:
                print('Ошибка. Повторите ввод.')
    # получение имён файлов
    im_files = list(map(find_im_names, path_list))[0]
    im_files = list(filter(lambda strok: strok.split('.')[1] == 'jpg' or strok.split('.')[1] == 'png', im_files))
    # преобразование параметров в массивы и списки
    if not k_step:
        k_list = [k_start]
    else:
        k_list = np.arange(k_start, k_finish, k_step)
    if not p_step:
        p_list = [p_start]
    else:
        p_list = np.arange(p_start, p_finish, p_step)
    if not non_max_size_step:
        non_max_size_list = [non_max_size_start]
    else:
        non_max_size_list = np.arange(non_max_size_start, non_max_size_finish, non_max_size_step)
    if not b_step:
        b_list = [b_start]
    else:
        b_list = np.arange(b_start, b_finish, b_step)
    if not cut_th_step:
        cut_th_list = [cut_th_start]
    else:
        cut_th_list = np.arange(cut_th_start, cut_th_finish, cut_th_step)

    print('По умолчанию результаты будут помещены в файл '
          'C://Users/Настя/Documents/Компьютерное зрение/Исследование детектора Харриса/Results.xlsx')
    print('Изменить имя файла (y(Y)/n(N))?')
    symb = input()
    if symb.lower() != 'n':
        print('Введите имя файла, в котором будут храниться результаты, в формате C://путь/к/файлу/название.xlsx')
        result_file = input()
    else:
        result_file = 'C://Users/Настя/Documents/Компьютерное зрение/Исследование детектора Харриса/Results.xlsx'
        
    # поиск углов
    tstart = time.time()
    with cf.ThreadPoolExecutor() as tpe:
        f = tpe.map(find_corners, im_files)
    tfinish = time.time()
    print('Сохраняю результаты')
    save_results()
    print('Программа завершена')
    print('Время на поиск углов, сек:', tfinish - tstart)
